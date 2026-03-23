"""
Detecta archivos {SIGLA}{AÑO}-{PERIODO}.xlsx → genera ALL_DATA en index.html
Uso: python build_data.py
"""
import json
import re
import sys
import glob as globmod
import openpyxl

HTML_PATH = "index.html"

FACULTY_MAP = {
    "FIEECS": {"label": "FIEECS \u00b7 UNI", "fullName": "Fac. Ing. El\u00e9ctrica y Electr\u00f3nica"},
    "FIGMM":  {"label": "FIGMM \u00b7 UNI",  "fullName": "Fac. Ing. Geol\u00f3gica, Minera y Metal\u00fargica"},
    "FIQT":   {"label": "FIQT \u00b7 UNI",   "fullName": "Fac. Ing. Qu\u00edmica y Textil"},
    "FIIS":   {"label": "FIIS \u00b7 UNI",   "fullName": "Fac. Ing. Industrial y Sistemas"},
    "FIEE":   {"label": "FIEE \u00b7 UNI",   "fullName": "Fac. Ing. El\u00e9ctrica y Electr\u00f3nica"},
    "FIPP":   {"label": "FIPP \u00b7 UNI",   "fullName": "Fac. Ing. Petr\u00f3leo, Gas Natural y Petroqu\u00edmica"},
    "FIM":    {"label": "FIM \u00b7 UNI",    "fullName": "Fac. Ing. Mec\u00e1nica"},
    "FIC":    {"label": "FIC \u00b7 UNI",    "fullName": "Fac. Ing. Civil"},
    "FIA":    {"label": "FIA \u00b7 UNI",    "fullName": "Fac. Ing. Ambiental"},
    "FC":     {"label": "FC \u00b7 UNI",     "fullName": "Fac. de Ciencias"},
}

# Mapeo sección → especialidad (FIA)
SECC_ESP = {
    "E": "IS",
    "F": "IH",
    "G": "IA",
    "H": "IS",
    "I": "CB",
    "J": "IA",
}

TIPO_MAP = {
    "TEORIA": "T", "TEORÍA": "T",
    "PRACTICA": "P", "PRÁCTICA": "P",
    "LABORATORIO": "L", "LAB": "L",
    "SEMINARIO": "S",
}


def parse_horario(h):
    """'MA 10-12' → ('MA', 10, 12)"""
    parts = str(h).strip().split()
    if len(parts) != 2:
        return None
    rng = parts[1].split("-")
    if len(rng) != 2:
        return None
    try:
        return parts[0].upper(), int(rng[0]), int(rng[1])
    except ValueError:
        return None


def get_ciclo(cod):
    c = str(cod)[2] if len(str(cod)) > 2 else ""
    return int(c) if c.isdigit() else 11


def load_rows(xlsx_path):
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    rows = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        all_rows = list(ws.iter_rows(values_only=True))

        hi = -1
        headers = []
        for i, row in enumerate(all_rows):
            cells = [str(c).upper().strip() if c is not None else "" for c in row]
            if "COD" in cells and "CURSO" in cells:
                hi = i
                headers = cells
                break
        if hi < 0:
            continue

        def ix(name):
            try:
                return headers.index(name)
            except ValueError:
                return -1

        iC = ix("COD")
        iCu = ix("CURSO")
        iS = ix("SECC")
        iTi = ix("TIPO")
        iH = ix("HORARIO")
        iD = ix("DIA")
        iHi = ix("H INI")
        iHf = ix("H FIN")
        iSa = ix("AULA") if ix("AULA") >= 0 else ix("SALON")
        iDo = ix("DOCENTE")
        iCy = ix("CICLO")
        fia_fmt = iH >= 0

        for row in all_rows[hi + 1:]:
            if not row[iC] or not row[iCu]:
                continue

            cod = str(row[iC]).strip()
            curso = str(row[iCu]).strip()
            secc = str(row[iS] or "").strip().upper() if iS >= 0 else ""
            tipo_raw = str(row[iTi] or "").strip().upper() if iTi >= 0 else ""
            tipo = TIPO_MAP.get(tipo_raw, "T")
            ciclo = int(row[iCy]) if iCy >= 0 and row[iCy] and str(row[iCy]).strip().isdigit() else get_ciclo(cod)
            salon = str(row[iSa] or "").strip() if iSa >= 0 else ""
            docente = str(row[iDo] or "").strip() if iDo >= 0 else ""
            esp = SECC_ESP.get(secc, "")

            if fia_fmt:
                ph = parse_horario(row[iH])
                if not ph:
                    continue
                dia, h0, h1 = ph
                rows.append({"esp": esp, "cod": cod, "secc": secc, "curso": curso,
                              "docente": docente, "tipo": tipo, "dia": dia,
                              "hIni": h0, "hFin": h1, "salon": salon})
            else:
                if iD < 0 or not row[iD]:
                    continue
                try:
                    h0, h1 = int(row[iHi]), int(row[iHf])
                except (TypeError, ValueError):
                    continue
                dia = str(row[iD]).strip().upper()
                rows.append({"esp": esp, "cod": cod, "secc": secc, "curso": curso,
                              "docente": docente, "tipo": tipo, "dia": dia,
                              "hIni": h0, "hFin": h1, "salon": salon})

    wb.close()
    return rows


def find_faculty_files():
    """Busca archivos {SIGLA}{AÑO}-{PERIODO}.xlsx en el directorio actual."""
    # Ordenar por longitud descendente para que FIEECS no sea confundido con FIA
    siglas = sorted(FACULTY_MAP.keys(), key=len, reverse=True)
    pattern = re.compile(
        r'^(' + '|'.join(re.escape(s) for s in siglas) + r')(\d{4}-\d)\.xlsx$'
    )
    found = {}
    for f in globmod.glob("*.xlsx"):
        m = pattern.match(f)
        if m:
            sigla = m.group(1)
            period = m.group(2)
            found[sigla] = {"file": f, "period": period}
    return found


def update_html(all_data):
    with open(HTML_PATH, "r", encoding="utf-8") as f:
        content = f.read()

    new_js = "const ALL_DATA=" + json.dumps(all_data, ensure_ascii=False, separators=(",", ":")) + ";"

    # Replace const FIA_DATA=[...]; (first deploy)
    new_content, n = re.subn(r"const FIA_DATA=\[.*?\];", new_js, content)
    if n == 0:
        # Replace existing ALL_DATA={...};
        new_content, n = re.subn(r"const ALL_DATA=\{.*?\};", new_js, content)
    if n == 0:
        print("ERROR: no se encontró 'const FIA_DATA=[...]' ni 'const ALL_DATA={...}' en index.html", file=sys.stderr)
        sys.exit(1)

    with open(HTML_PATH, "w", encoding="utf-8") as f:
        f.write(new_content)

    siglas = list(all_data.keys())
    total = sum(len(v["rows"]) for v in all_data.values())
    print(f"OK: {len(siglas)} facultad(es) {siglas}, {total} sesiones totales")


if __name__ == "__main__":
    found = find_faculty_files()
    if not found:
        print("ERROR: no se encontró ningún archivo {SIGLA}{AÑO}-{PERIODO}.xlsx", file=sys.stderr)
        sys.exit(1)

    all_data = {}
    for sigla, info in found.items():
        rows = load_rows(info["file"])
        if not rows:
            print(f"WARN: {info['file']} sin filas, omitido")
            continue
        meta = FACULTY_MAP[sigla]
        all_data[sigla] = {
            "label": meta["label"],
            "fullName": meta["fullName"],
            "period": info["period"],
            "rows": rows,
        }

    if not all_data:
        print("ERROR: ningún xlsx con datos válidos", file=sys.stderr)
        sys.exit(1)

    update_html(all_data)
